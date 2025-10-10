import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, time
import os
import string
import warnings
import unicodedata
import tempfile
import re
import openpyxl
from openpyxl.styles import PatternFill
import math
from io import BytesIO
import json

# Suprime avisos específicos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

# Carregar configurações do arquivo JSON
def load_config():
    try:
        with open("config.json", "r", encoding="utf-8") as f:
            config = json.load(f)
        return (
            config["required_columns"],
            config["buyer_carrossel_map"],
            config["product_name_corrections"]
        )
    except FileNotFoundError:
        st.error("Arquivo config.json não encontrado. Certifique-se de que ele está no mesmo diretório do script.")
        return None, None, None
    except Exception as e:
        st.error(f"Erro ao carregar config.json: {e}")
        return None, None, None

# Carregar as configurações
required_columns, buyer_carrossel_map, product_name_corrections = load_config()

# Verificar se as configurações foram carregadas corretamente
if required_columns is None or buyer_carrossel_map is None or product_name_corrections is None:
    st.stop()

# Função para corrigir códigos interpretados como datas
def fix_if_date(value):
    if pd.isna(value):
        return value
    if isinstance(value, (datetime, pd.Timestamp)):
        return f"{value.year}-{value.month}"
    else:
        return str(value)

def get_unique_filename(path):
    """Recebe um caminho de arquivo e retorna um nome único no mesmo diretório."""
    base, ext = os.path.splitext(path)
    counter = 1
    new_path = path
    while os.path.exists(new_path):
        new_path = f"{base} ({counter}){ext}"
        counter += 1
    return new_path

# --- Função para limpar e converter preços ---
def clean_price_value(value):
    """Limpa string de preço e converte para float, retornando None se inválido."""
    try:
        text = str(value).replace("R$", "").replace(",", ".").strip()
        return float(text)
    except:
        return None

def detect_header_with_scoring(df):
    """
    Sistema de pontuação para descobrir qual linha é o verdadeiro header.
    Retorna a linha que será usada como header e lista de erros caso alguma coluna obrigatória não seja encontrada.
    """
    max_score = -1
    header_row = None
    row_found = None

    # Limitar iteração às primeiras 20 linhas
    limited_df = df.head(20)
    
    for idx, row in limited_df.iterrows():
        score = 0
        normalized_row = [normalize_text(str(cell)) for cell in row if not pd.isna(cell)]

        for col in required_columns:
            if normalize_text(col) in normalized_row:
                score += 1

        if score > max_score:
            max_score = score
            header_row = idx
            row_found = normalized_row

    # Verificar se nenhuma linha válida foi encontrada
    if header_row is None or not row_found:
        errors = ["❌ Nenhuma linha válida encontrada com colunas obrigatórias nas primeiras 20 linhas. Verifique o arquivo."]
        return None, errors

    # Verificar se todas as obrigatórias estão presentes
    missing_cols = [col for col in required_columns if normalize_text(col) not in row_found]
    if missing_cols:
        errors = [
            f"❌ Coluna obrigatória '{col}' não encontrada na linha {header_row + 1} do arquivo do Encarte Consolidado. Verifique a digitação do título da coluna."
            for col in missing_cols
        ]
        return None, errors

    return header_row, []

# --- Função para normalização de texto para matching ---
translator = str.maketrans('', '', string.punctuation)
def normalize_text(text):
    """Normaliza texto: lowercase, remove acentos e pontuação."""
    if pd.isna(text):
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
    text = text.translate(translator)
    return text

def get_carrossel_value(normalized_buyer, mapping):
    """Retorna o valor do carrossel se chave estiver contida no texto do comprador."""
    if not normalized_buyer:
        return ''
    for key, value in mapping.items():
        if key in normalized_buyer:
            return value
    return ''

def correct_product_name(name):
    """Corrige o nome do produto com base no dicionário de correções, retornando em maiúsculo."""
    if pd.isna(name):
        return ""
    corrected_name = str(name).strip()
    for pattern, replacement in product_name_corrections.items():
        corrected_name = re.sub(pattern, replacement, corrected_name, flags=re.IGNORECASE)
    return corrected_name.upper()

def remove_suffix(text):
    """Remove sufixos como _sell out, _faturamento e tudo que vier depois."""
    if pd.isna(text):
        return ""
    keywords = ['sell out', 'faturamento', 'sell in']
    pattern = r'_(' + '|'.join(keywords) + r').*$'
    return re.sub(pattern, '', str(text), flags=re.IGNORECASE).strip()

def classify_ean(ean_str):
    """
    Classifica o EAN retornando uma tupla (tipo_codigo, unidade).
    Regras:
      - Se tinha "/" originalmente → Interno, Quilograma
      - Se todos os códigos < 12 dígitos → Interno, Quilograma
      - Caso contrário → EAN, Unidade
    """
    if not ean_str or pd.isna(ean_str) or not str(ean_str).strip():
        return ("EAN", "Unidade")

    ean_str = str(ean_str)
    had_slash = "/" in ean_str
    ean_str = ean_str.replace("/", ";")
    eans = [e.strip() for e in ean_str.split(';') if e.strip()]
    if not eans:
        return ("EAN", "Unidade")
    
    first_ean = eans[0]
    first_len = len(first_ean)

    if had_slash or first_len < 12:
        return ("Interno", "Quilograma")
    else:
        return ("EAN", "Unidade")

def get_code_type(ean):
    if pd.isna(ean) or not str(ean).strip():
        return 'EAN'
    ean_str = str(ean)
    if "/" in ean_str:
        return 'Interno'
    eans = [e.strip() for e in ean_str.split(';') if e.strip()]
    if not eans:
        return 'EAN'
    lens = [len(e) for e in eans]
    if all(l < 12 for l in lens):
        return 'Interno'
    else:
        return 'EAN'

# Função principal de montagem do DataFrame
def build_final_dataframe(filtered_df, profile, start_date, end_date, store_map, apply_name_correction):
    df_copy = filtered_df.copy()
    df_copy['descrição do item'] = df_copy['descrição do item'].apply(remove_suffix)

    if apply_name_correction:
        df_copy['descrição do item'] = df_copy['descrição do item'].apply(correct_product_name)
    else:
        df_copy['descrição do item'] = df_copy['descrição do item'].apply(
            lambda x: str(x).strip().upper() if not pd.isna(x) else ""
        )

    if df_copy.empty:
        st.warning(f"Nenhuma linha válida encontrada para o perfil {profile}. O arquivo não será gerado.")
        return None

    df_copy['ean'] = df_copy['ean'].astype(str).str.replace("/", ";", regex=False)
    df_copy[['final_code_type', 'final_unit']] = df_copy['ean_original_encarte'].apply(
        lambda x: pd.Series(classify_ean(x))
    )

    possible_buyer_names = ['comprador', 'compradora', 'compradores', 'compradoras']
    col_name = next((col for col in possible_buyer_names if col in df_copy.columns), None)

    if col_name:
        df_copy['comprador_normalized'] = df_copy[col_name].apply(normalize_text)
    else:
        df_copy['comprador_normalized'] = ''
        st.warning("Nenhuma coluna de comprador encontrada. 'Carrossel' ficará vazio.")

    df_copy['final_carrossel'] = df_copy['comprador_normalized'].apply(
        lambda x: get_carrossel_value(x, buyer_carrossel_map)
    )

    return pd.DataFrame({
        "Nome": df_copy["descrição do item"],
        "Carrossel": df_copy["final_carrossel"],
        "Check-In": "Não",
        "Preço": df_copy["preço de:"],
        "Preço promocional": df_copy["preço por:"],
        "Limite por cliente": 0,
        "Dias para Resgate após ativação": (end_date.date() - start_date.date()).days + 1,
        "Unidade": df_copy["final_unit"],
        "Não exigir ativação no App": "Ativação automática",
        "Ativar em": start_date.strftime("%d/%m/%Y %H:%M"),
        "Inativar em": end_date.strftime("%d/%m/%Y %H:%M"),
        "URL da imagem": "",
        "Tipo do código": df_copy["final_code_type"],
        "Códigos dos produtos": df_copy["ean"],
        "Tipo Promocional": "De / por",
        "Sobrescrever lojas": "Sim",
        "Lojas": store_map[profile]
    })

# --- Função para mesclar EANs do arquivo ---
def merge_ean_data(df_base, ean_file):
    try:
        file_extension = os.path.splitext(ean_file.name)[1].lower()
        if file_extension in ['.xlsx', '.xls']:
            df_ean = pd.read_excel(ean_file)
        elif file_extension == '.csv':
            df_ean = pd.read_csv(ean_file, sep=';')
        else:
            st.error("Formato de arquivo de EANs não suportado. Use xlsx, xls ou csv.")
            return df_base
        
        df_ean = df_ean.rename(columns={'CÓDIGO PRODUTO': 'código', 'CÓDIGO EAN': 'ean'})
        if 'código' in df_ean.columns:
            df_ean['código'] = df_ean['código'].apply(fix_if_date)
        if 'ean' in df_ean.columns:
            df_ean['ean'] = df_ean['ean'].apply(fix_if_date)
        
        df_base['código'] = df_base['código'].astype(str).str.strip().str.replace('-', '')
        df_ean['código'] = df_ean['código'].astype(str).str.strip().str.replace('-', '')
        
        expanded_rows = []
        for _, row in df_base.iterrows():
            new_row = row.copy()
            encarte_ean = str(new_row['ean']).strip() if not pd.isna(new_row['ean']) else ""
            encarte_ean = encarte_ean.replace('/', ';')
            encarte_ean_list = [e.strip() for e in encarte_ean.split(';') if e.strip() and e != 'nan']
            
            matching_eans = df_ean[df_ean['código'] == new_row['código']]['ean']
            
            if not matching_eans.empty:
                ean_list = []
                for ean in matching_eans:
                    if not pd.isna(ean) and str(ean).strip() and str(ean).strip() != 'nan':
                        eans = str(ean).strip().replace('/', ';').split(';')
                        ean_list.extend([e.strip() for e in eans if e.strip()])
                ean_list.extend(encarte_ean_list)
                ean_list = list(dict.fromkeys(ean_list))
                new_row['ean'] = ';'.join(ean_list) if ean_list else encarte_ean
            else:
                new_row['ean'] = ';'.join(encarte_ean_list) if encarte_ean_list else encarte_ean
            
            expanded_rows.append(new_row)
        
        df_updated = pd.DataFrame(expanded_rows)
        return df_updated
    except Exception as e:
        st.error(f"Erro ao mesclar dados de EAN: {e}")
        return df_base

# --- Função para listar planilhas disponíveis ---
def list_sheets(uploaded_file):
    file_extension = os.path.splitext(uploaded_file.name)[1].lower()
    try:
        if file_extension in ['.xlsx', '.xls']:
            xl = pd.ExcelFile(uploaded_file)
            return xl.sheet_names
        elif file_extension == '.csv':
            return ["Planilha CSV"]
        else:
            st.error("Formato de arquivo não suportado. Use xlsx, xls ou csv.")
            return []
    except Exception as e:
        st.error(f"Erro ao listar planilhas: {e}")
        return []

# --- Função principal para processar a planilha ---
def process_promotions(uploaded_file, ean_file, start_date, end_date, temp_dir, use_ean_file, apply_name_correction, sheet_name):
    profiles = ["GERAL/PREMIUM", "GERAL", "PREMIUM"]
    store_mapping = {
        "GERAL": "4368-4363-4362-4357-4360-4356-4370-4359-4372-4353-4371-4365-4369-4361-4366-4354-4355-4364",
        "PREMIUM": "4373-4358-4367",
        "GERAL/PREMIUM": "4368-4363-4362-4357-4360-4356-4370-4359-4372-4353-4371-4365-4369-4361-4366-4354-4355-4364-4373-4358-4367"
    }

    file_extension = os.path.splitext(uploaded_file.name)[1].lower()
    try:
        if file_extension in ['.xlsx', '.xls']:
            temp_df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
        elif file_extension == '.csv':
            temp_df = pd.read_csv(uploaded_file, sep=';', header=None)
        else:
            st.error("Formato de arquivo base não suportado. Use xlsx, xls ou csv.")
            return []
    except Exception as e:
        st.error(f"Erro ao ler o arquivo base: {e}")
        return []

    header_row, errors = detect_header_with_scoring(temp_df)
    if errors:
        for msg in errors:
            st.error(msg)
        return []

    try:
        if file_extension in ['.xlsx', '.xls']:
            df_base = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row)
        else:
            df_base = pd.read_csv(uploaded_file, sep=';', header=header_row)
    except Exception as e:
        st.error(f"Erro ao ler o arquivo com o cabeçalho detectado: {e}")
        return []

    df_base.columns = df_base.columns.str.strip().str.replace(r'\s+', ' ', regex=True).str.lower()
    
    if 'código' in df_base.columns:
        df_base['código'] = df_base['código'].apply(fix_if_date)
    if 'ean' in df_base.columns:
        df_base['ean'] = df_base['ean'].apply(fix_if_date)  

    df_base['ean_original_encarte'] = df_base['ean']
    df_base["preço de:"] = df_base["preço de:"].apply(clean_price_value)
    df_base["preço por:"] = df_base["preço por:"].apply(clean_price_value)
    df_base["copied_preço_de"] = False
    df_base["copied_preço_por"] = False

    for i in range(len(df_base)):
        if pd.isna(df_base.iloc[i]["preço de:"]) and i > 0:
            current_ean = str(df_base.iloc[i]["ean"]) if not pd.isna(df_base.iloc[i]["ean"]) else ""
            prev_ean = str(df_base.iloc[i-1]["ean"]) if not pd.isna(df_base.iloc[i-1]["ean"]) else ""
            if current_ean[:7] == prev_ean[:7] and not pd.isna(df_base.iloc[i-1]["preço de:"]):
                df_base.iloc[i, df_base.columns.get_loc("preço de:")] = df_base.iloc[i-1]["preço de:"]
                df_base.iloc[i, df_base.columns.get_loc("copied_preço_de")] = True
        
        if pd.isna(df_base.iloc[i]["preço por:"]) and i > 0:
            current_ean = str(df_base.iloc[i]["ean"]) if not pd.isna(df_base.iloc[i]["ean"]) else ""
            prev_ean = str(df_base.iloc[i-1]["ean"]) if not pd.isna(df_base.iloc[i-1]["ean"]) else ""
            if current_ean[:7] == prev_ean[:7] and not pd.isna(df_base.iloc[i-1]["preço por:"]):
                df_base.iloc[i, df_base.columns.get_loc("preço por:")] = df_base.iloc[i-1]["preço por:"]
                df_base.iloc[i, df_base.columns.get_loc("copied_preço_por")] = True

    if use_ean_file and ean_file:
        df_base = merge_ean_data(df_base, ean_file)

    df_base['perfil de loja'] = df_base['perfil de loja'].ffill()
    df_base['tipo ação'] = df_base['tipo ação'].ffill()
    df_filtered = df_base[df_base["tipo ação"].str.contains("CRM", case=False, na=False)]

    os.makedirs(temp_dir, exist_ok=True)
    output_files = []

    for profile in profiles:
        df_profile = df_filtered[df_filtered["perfil de loja"] == profile].copy()
        if df_profile.empty:
            st.warning(f"Nenhuma linha encontrada para o perfil {profile}. Pulando geração.")
            continue
        
        df_final = build_final_dataframe(df_profile, profile, start_date, end_date, store_mapping, apply_name_correction)
        if df_final is None or df_final.empty:
            st.warning(f"O DataFrame final do perfil {profile} está vazio. Pulando exportação.")
            continue

        filename = f"promo_{profile.replace('/', '_')}_CRM.xlsx"
        filepath = os.path.join(temp_dir, filename)
        filepath = get_unique_filename(filepath)

        df_final.to_excel(filepath, index=False, engine="openpyxl")
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        header_values = [cell.value for cell in ws[1]]
        preco_col = header_values.index("Preço") + 1
        preco_promo_col = header_values.index("Preço promocional") + 1
        unidade_col = header_values.index("Unidade") + 1
        tipo_codigo_col = header_values.index("Tipo do código") + 1

        for row_idx in range(2, ws.max_row + 1):
            preco_cell = ws.cell(row=row_idx, column=preco_col)
            preco_promo_cell = ws.cell(row=row_idx, column=preco_promo_col)
            unidade_cell = ws.cell(row=row_idx, column=unidade_col)
            tipo_codigo_cell = ws.cell(row=row_idx, column=tipo_codigo_col)

            if preco_cell.value is None or str(preco_cell.value).strip() in ("", "nan") or (
                isinstance(preco_cell.value, float) and math.isnan(preco_cell.value)
            ):
                preco_cell.fill = red_fill

            if preco_promo_cell.value is None or str(preco_promo_cell.value).strip() in ("", "nan") or (
                isinstance(preco_promo_cell.value, float) and math.isnan(preco_promo_cell.value)
            ):
                preco_promo_cell.fill = red_fill

            if str(unidade_cell.value).strip().upper() == "QUILOGRAMA":
                unidade_cell.fill = yellow_fill

            if str(tipo_codigo_cell.value).strip().upper() == "INTERNO":
                tipo_codigo_cell.fill = yellow_fill

        wb.save(filepath)

        with open(filepath, "rb") as f:
            output = BytesIO(f.read())
            output.seek(0)

        output_files.append((filename, output))
        st.success(f"✅ Arquivo gerado: {filename}")

    return output_files

# --- Interface Streamlit ---
st.title("Processador de Promoções CRM")
st.write("Faça upload da planilha de promoções (xlsx, xls ou csv) e, opcionalmente, um arquivo com EANs (xlsx, xls ou csv). Selecione as datas do encarte e a planilha desejada.")

temp_dir = tempfile.mkdtemp()
default_start = datetime.today()
default_end = datetime.today() + timedelta(days=7)

col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input("Data de Início do Encarte", value=default_start, format="DD/MM/YYYY")
with col2:
    end_date = st.date_input("Data de Fim do Encarte", value=default_end, format="DD/MM/YYYY")

if end_date < start_date:
    st.error("A data de fim não pode ser anterior à data de início.")
else:
    apply_name_correction = st.checkbox("Aplicar correção de nomes de produtos", value=False)
    use_ean_file = st.checkbox("Usar arquivo de EANs", value=False)
    uploaded_file = st.file_uploader("Selecione o arquivo de ENCARTE CONSOLIDADO", type=["xlsx", "xls", "csv"])
    
    selected_sheet = None
    if uploaded_file:
        sheet_names = list_sheets(uploaded_file)
        if sheet_names:
            st.write("Selecione a planilha para processar:")
            selected_sheet = st.selectbox("Planilhas disponíveis", sheet_names)
        else:
            st.error("Nenhuma planilha encontrada no arquivo.")
    
    ean_file = None
    if use_ean_file:
        ean_file = st.file_uploader("Selecione o arquivo de EANs (opcional)", type=["xlsx", "xls", "csv"])

    if uploaded_file and selected_sheet:
        if st.button("Processar Promoções"):
            output_files = []
            try:
                with st.spinner("Processando..."):
                    start_date = datetime.combine(start_date, time(0, 0))
                    end_date = datetime.combine(end_date, time(23, 59))
                    output_files = process_promotions(uploaded_file, ean_file, start_date, end_date, temp_dir, use_ean_file, apply_name_correction, selected_sheet)
            except Exception as e:
                st.error(f"Erro durante o processamento: {e}")
            if output_files:
                for filename, output in output_files:
                    st.download_button(
                        label=f"Baixar {filename}",
                        data=output,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.warning("Nenhum arquivo foi gerado. Verifique os dados de entrada.")