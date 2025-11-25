import math
import openpyxl
from openpyxl.styles import PatternFill

def export_to_excel(df_final, filepath):
    """Exporta DataFrame para Excel com formatação especial"""
    df_final.to_excel(filepath, index=False, engine="openpyxl")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Aplicar formato de texto à coluna "Códigos dos produtos"
    header_values = [cell.value for cell in ws[1]]
    ean_col = header_values.index("Códigos dos produtos") + 1

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=ean_col).number_format = '@'

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Identificação de colunas existentes
    header_values = [cell.value for cell in ws[1]]
    preco_col = header_values.index("Preço") + 1
    preco_promo_col = header_values.index("Preço promocional") + 1
    unidade_col = header_values.index("Unidade") + 1
    tipo_codigo_col = header_values.index("Tipo do código") + 1

    for row_idx in range(2, ws.max_row + 1):

        # ========== EAN EM VERMELHO ==========
        ean_cell = ws.cell(row=row_idx, column=ean_col)
        ean_val = ean_cell.value

        if (
            ean_val is None
            or str(ean_val).strip() == ""
            or str(ean_val).strip().lower() == "nan"
            or (isinstance(ean_val, float) and math.isnan(ean_val))
        ):
            ean_cell.fill = red_fill

        # ---------- LÓGICA ORIGINAL (NADA FOI EXCLUÍDO) ----------
        preco_cell = ws.cell(row=row_idx, column=preco_col)
        preco_promo_cell = ws.cell(row=row_idx, column=preco_promo_col)
        unidade_cell = ws.cell(row=row_idx, column=unidade_col)
        tipo_codigo_cell = ws.cell(row=row_idx, column=tipo_codigo_col)

        if preco_cell.value is None or str(preco_cell.value).strip() in ("", "nan") or (
            isinstance(preco_cell.value, float) and math.isnan(preco_cell.value)
        ):
            preco_cell.fill = red_fill

        if preco_promo_cell.value is None or str(preco_promo_cell.value).strip() in ("", "nan") or (
            isinstance(preco_promo_cell.value, float) and math.isnan(preco_promo_cell.value)
        ):
            preco_promo_cell.fill = red_fill

        if str(unidade_cell.value).strip().upper() == "QUILOGRAMA":
            unidade_cell.fill = yellow_fill

        if str(tipo_codigo_cell.value).strip().upper() == "INTERNO":
            tipo_codigo_cell.fill = yellow_fill

    wb.save(filepath)
